# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('C:\\Users\\Admin\\Desktop\\excel.xlsx')

# create the sheet object
sheet = wb.active


def excel():
	
	# resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "Name"
	sheet.cell(row=1, column=2).value = "Course"
	sheet.cell(row=1, column=3).value = "Semester"
	sheet.cell(row=1, column=4).value = "Form Number"
	sheet.cell(row=1, column=5).value = "Contact Number"
	sheet.cell(row=1, column=6).value = "Email id"
	sheet.cell(row=1, column=7).value = "Address"


# Function to set focus (cursor)
def focus1(event):
	# set focus on the course_field box
	course_field.focus_set()


# Function to set focus
def focus2(event):
	# set focus on the sem_field box
	sem_field.focus_set()


# Function to set focus
def focus3(event):
	# set focus on the form_no_field box
	form_no_field.focus_set()


# Function to set focus
def focus4(event):
	# set focus on the contact_no_field box
	contact_no_field.focus_set()


# Function to set focus
def focus5(event):
	# set focus on the email_id_field box
	email_id_field.focus_set()


# Function to set focus
def focus6(event):
	# set focus on the address_field box
	address_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
	
	# clear the content of text entry box
	name_field.delete(0, END)
	course_field.delete(0, END)
	sem_field.delete(0, END)
	form_no_field.delete(0, END)
	contact_no_field.delete(0, END)
	email_id_field.delete(0, END)
	address_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
	
	# if user not fill any entry
	# then print "empty input"
	if (name_field.get() == "" and
		course_field.get() == "" and
		sem_field.get() == "" and
		form_no_field.get() == "" and
		contact_no_field.get() == "" and
		email_id_field.get() == "" and
		address_field.get() == ""):
			
		print("empty input")

	else:

		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
		current_row = sheet.max_row
		current_column = sheet.max_column

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
		sheet.cell(row=current_row + 1, column=1).value = name_field.get()
		sheet.cell(row=current_row + 1, column=2).value = course_field.get()
		sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
		sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
		sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
		sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
		sheet.cell(row=current_row + 1, column=7).value = address_field.get()

		# save the file
		wb.save('C:\\Users\\Admin\\Desktop\\excel.xlsx')

		# set focus on the name_field box
		name_field.focus_set()

		# call the clear() function
		clear()


# Driver code
if __name__ == "__main__":
    root = Tk()
    root.configure(background='light blue')
    root.title("Registration Form")
    root.geometry("600x350")

    excel()

    heading = Label(root, text="Student Registration Form", bg="light blue", font=("Helvetica", 20))

    labels = ["Name", "Course", "Semester", "Form Number", "Contact Number", "Email id", "Address"]
    for i, label_text in enumerate(labels):
        label = Label(root, text=label_text, bg="light blue")
        label.grid(row=i+1, column=0, padx=10, pady=10, sticky="w")

    name_field = Entry(root, font=("Helvetica", 15))
    course_field = Entry(root, font=("Helvetica", 15))
    sem_field = Entry(root, font=("Helvetica", 15))
    form_no_field = Entry(root, font=("Helvetica", 15))
    contact_no_field = Entry(root, font=("Helvetica", 15))
    email_id_field = Entry(root, font=("Helvetica", 15))
    address_field = Entry(root, font=("Helvetica", 15))

    fields = [name_field, course_field, sem_field, form_no_field, contact_no_field, email_id_field, address_field]

    for i, field in enumerate(fields):
        field.grid(row=i+1, column=1, ipadx="100", padx=10, pady=10)

    for i in range(7):
        fields[i].bind("<Return>", lambda event, index=i: fields[index+1].focus_set())

    submit = Button(root, text="Submit", fg="black", bg="#3085C3", font=("Helvetica", 14), command=insert)
    submit.grid(row=8, column=1, pady=20)

    root.mainloop()